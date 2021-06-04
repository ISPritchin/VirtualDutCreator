import pandas as pd
import openpyxl

from DataChecker import DataChecker
from TableTransformer import TableTransformer
from TableSplitter import TableSplitter


class FuelSystemConfigCreator:

    def __init__(self, path):
        self.path = path
        self.fuel_system_config = {}
        self.open_file()
        self.get_virtual_dut_table()

    def open_file(self):
        try:
            wb = openpyxl.load_workbook(self.path)
        except Exception:
            raise Exception(f"open error file {self.path}.")
        n_sheets = len(wb.sheetnames)
        for sheet_number in range(n_sheets):
            tank_name = wb.sheetnames[sheet_number]
            self.fuel_system_config[tank_name] = {}
            calibration_table = pd.read_excel(self.path, sheet_name=sheet_number, engine='openpyxl')
            DataChecker(calibration_table).check_table()
            self.fuel_system_config[tank_name]["calibration_table"] = calibration_table

    def get_virtual_dut_table(self):
        for tank_name in self.fuel_system_config:
            calibration_table = self.fuel_system_config[tank_name]["calibration_table"]
            _ = TableTransformer(calibration_table).transform()
            self.fuel_system_config[tank_name]["virtual_duts_table"] = TableSplitter(_).split()

    def create(self):
        self.open_file()
        self.get_virtual_dut_table()
        return self.fuel_system_config


# if __name__ == "__main__":
#     FSC = FuelSystemConfigCreator(r"test/configurations/HardConfiguration/test1.xlsx").create()
#     vduts = FSC["Tank1"]["virtual_duts_table"]
#     for x in vduts:
#         print(x)